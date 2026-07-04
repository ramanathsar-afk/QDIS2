from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


class Report:

    def __init__(self):

        self.folder = Path("reports")
        self.folder.mkdir(exist_ok=True)

    def save(self, df):

        if df.empty:
            print("No report.")
            return None

        df = df.sort_values(
            "Overall",
            ascending=False
        ).reset_index(drop=True)

        df.insert(0, "Rank", range(1, len(df)+1))

        filename = self.folder / (
            "QDIS_Report_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".xlsx"
        )

        with pd.ExcelWriter(
            filename,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Ranking",
                index=False
            )

            ws = writer.sheets["Ranking"]

            # Header

            fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            for cell in ws[1]:

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = fill

                cell.alignment = Alignment(
                    horizontal="center"
                )

            # Column widths

            for col in ws.columns:

                width = max(
                    len(str(c.value))
                    if c.value is not None else 0
                    for c in col
                )

                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = min(width + 3, 35)

            # Freeze header

            ws.freeze_panes = "A2"

            # Auto filter

            ws.auto_filter.ref = ws.dimensions

            # Conditional colours

            for r in range(2, ws.max_row + 1):

                rating = ws[f"G{r}"].value

                if rating == "STRONG BUY":

                    colour = "00B050"

                elif rating == "BUY":

                    colour = "92D050"

                elif rating == "WATCH":

                    colour = "FFD966"

                else:

                    colour = "F4CCCC"

                ws[f"G{r}"].fill = PatternFill(
                    fill_type="solid",
                    fgColor=colour
                )

        print(f"\nReport saved : {filename}")

        return filename